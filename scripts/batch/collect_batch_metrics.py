#!/usr/bin/env python
"""
Collect key metrics from all batch results.

Extracts:
- Time to 39mm Gape
- Force at 39mm Gape
- Time to failure (max force)
- Gape at Failure
- Initial Gape (if delta calculated)
- Delta Gape at Failure (if delta calculated)

Usage (from project root):
    python scripts/batch/collect_batch_metrics.py
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np
import re

# Add project root to Python path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.calibration import load_calibration


def natural_sort_key(s):
    """Sort key for natural ordering (A1, A2, ..., A10, A11, ...)"""
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s)]


def find_time_to_gape(df, target_total_gape_mm, pixels_per_mm):
    """
    Find the time when gape first reaches target value using delta gape.

    Args:
        df: DataFrame with Time, Delta_Gape_px, and Initial_Gape_px columns
        target_total_gape_mm: Target total gape in mm (e.g., 39mm)
        pixels_per_mm: Calibration factor

    Returns:
        Time in seconds, or None if not reached
    """
    # Check if Delta_Gape_px column exists
    if 'Delta_Gape_px' not in df.columns or 'Initial_Gape_px' not in df.columns:
        return None

    # Get initial gape in mm
    initial_gape_mm = df['Initial_Gape_px'].iloc[0] / pixels_per_mm

    # Calculate target delta gape: delta = total - initial
    target_delta_mm = target_total_gape_mm - initial_gape_mm

    # Convert delta gape from pixels to mm
    df['Delta_Gape_mm'] = df['Delta_Gape_px'] / pixels_per_mm

    # Find first row where delta gape >= target
    matches = df[df['Delta_Gape_mm'] >= target_delta_mm]

    if len(matches) == 0:
        return None

    return matches.iloc[0]['Time']


def find_force_at_gape(df, target_total_gape_mm, pixels_per_mm):
    """
    Find the force when gape first reaches target value using delta gape.

    Args:
        df: DataFrame with Time, Force, Delta_Gape_px, and Initial_Gape_px columns
        target_total_gape_mm: Target total gape in mm (e.g., 39mm)
        pixels_per_mm: Calibration factor

    Returns:
        Force in N, or None if not reached
    """
    # Check if Delta_Gape_px column exists
    if 'Delta_Gape_px' not in df.columns or 'Initial_Gape_px' not in df.columns:
        return None

    # Get initial gape in mm
    initial_gape_mm = df['Initial_Gape_px'].iloc[0] / pixels_per_mm

    # Calculate target delta gape: delta = total - initial
    target_delta_mm = target_total_gape_mm - initial_gape_mm

    # Convert delta gape from pixels to mm
    df['Delta_Gape_mm'] = df['Delta_Gape_px'] / pixels_per_mm

    # Find first row where delta gape >= target
    matches = df[df['Delta_Gape_mm'] >= target_delta_mm]

    if len(matches) == 0:
        return None

    return matches.iloc[0]['Force']


def find_failure_metrics(df, pixels_per_mm):
    """
    Find time and gape at failure (maximum force).
    
    Args:
        df: DataFrame with Time, Force, and Gape_Distance_px columns
        pixels_per_mm: Calibration factor
    
    Returns:
        (time_at_failure, gape_at_failure_mm)
    """
    # Find row with maximum force
    max_force_idx = df['Force'].idxmax()
    max_force_row = df.loc[max_force_idx]
    
    time_at_failure = max_force_row['Time']
    gape_at_failure_mm = max_force_row['Gape_Distance_px'] / pixels_per_mm
    
    return time_at_failure, gape_at_failure_mm


def collect_sample_metrics(csv_path, sample_name, pixels_per_mm, target_gape_mm=39.0):
    """
    Collect metrics from a single sample.

    Args:
        csv_path: Path to synchronized CSV file
        sample_name: Sample identifier (e.g., "A1")
        pixels_per_mm: Calibration factor
        target_gape_mm: Target gape threshold in mm

    Returns:
        Dictionary with metrics
    """
    try:
        # Load synchronized data
        df = pd.read_csv(csv_path)

        # Calculate absolute gape metrics
        time_to_39mm = find_time_to_gape(df, target_gape_mm, pixels_per_mm)
        force_at_39mm = find_force_at_gape(df, target_gape_mm, pixels_per_mm)
        time_at_failure, gape_at_failure = find_failure_metrics(df, pixels_per_mm)

        metrics = {
            'Sample': sample_name,
            'Time_to_39mm_Gape_s': time_to_39mm,
            'Force_at_39mm_Gape_N': force_at_39mm,
            'Time_to_Failure_s': time_at_failure,
            'Gape_at_Failure_mm': gape_at_failure
        }

        # Check for delta gape columns and add delta metrics if available
        if 'Delta_Gape_px' in df.columns and df['Delta_Gape_px'].notna().any():
            initial_gape_px = df['Initial_Gape_px'].iloc[0] if 'Initial_Gape_px' in df.columns else np.nan
            initial_gape_mm = initial_gape_px / pixels_per_mm if not np.isnan(initial_gape_px) else np.nan

            # Max delta gape at failure
            max_force_idx = df['Force'].idxmax()
            delta_gape_at_failure_px = df.loc[max_force_idx, 'Delta_Gape_px']
            delta_gape_at_failure_mm = delta_gape_at_failure_px / pixels_per_mm

            metrics.update({
                'Initial_Gape_mm': initial_gape_mm,
                'Delta_Gape_at_Failure_mm': delta_gape_at_failure_mm
            })

        return metrics

    except Exception as e:
        print(f"Warning: Error processing {sample_name}: {e}")
        base_metrics = {
            'Sample': sample_name,
            'Time_to_39mm_Gape_s': None,
            'Force_at_39mm_Gape_N': None,
            'Time_to_Failure_s': None,
            'Gape_at_Failure_mm': None
        }
        # Try to determine if delta columns should be present
        try:
            df = pd.read_csv(csv_path)
            if 'Delta_Gape_px' in df.columns:
                base_metrics.update({
                    'Initial_Gape_mm': None,
                    'Delta_Gape_at_Failure_mm': None
                })
        except:
            pass
        return base_metrics


def main():
    print("=" * 70)
    print("Collecting Batch Results Metrics")
    print("=" * 70)

    # Define paths
    batch_results_dir = Path("results/batch")
    output_csv = Path("results/batch/batch_metrics_summary.csv")

    # Load calibration
    pixels_per_mm = load_calibration()
    print(f"\nUsing calibration: {pixels_per_mm:.3f} pixels/mm")
    print(f"Target gape threshold: 39.0 mm\n")

    # Generate list of all expected samples (A1-A25, B1-B25, ..., H1-H25)
    all_expected_samples = []
    for batch_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for sample_num in range(1, 26):
            all_expected_samples.append(f"{batch_letter}{sample_num}")

    # Create metrics dictionary with all expected samples
    metrics_dict = {}
    for sample in all_expected_samples:
        metrics_dict[sample] = {
            'Sample': sample,
            'Time_to_39mm_Gape_s': None,
            'Force_at_39mm_Gape_N': None,
            'Time_to_Failure_s': None,
            'Gape_at_Failure_mm': None,
            'Initial_Gape_mm': None,
            'Delta_Gape_at_Failure_mm': None
        }

    # Collect all batch directories
    batch_dirs = sorted([d for d in batch_results_dir.iterdir() if d.is_dir()],
                       key=lambda x: x.name)

    processed_count = 0

    # Process each batch
    for batch_dir in batch_dirs:
        batch_name = batch_dir.name
        print(f"Processing {batch_name}...")

        # Get all sample directories
        sample_dirs = sorted([d for d in batch_dir.iterdir() if d.is_dir()],
                           key=lambda x: natural_sort_key(x.name))

        for sample_dir in sample_dirs:
            sample_name = sample_dir.name
            csv_file = sample_dir / f"{sample_name}_synchronized.csv"

            if csv_file.exists():
                metrics = collect_sample_metrics(csv_file, sample_name, pixels_per_mm)
                if sample_name in metrics_dict:
                    metrics_dict[sample_name] = metrics
                    processed_count += 1
                    print(f"  ✓ {sample_name}")
            else:
                print(f"  ⚠ {sample_name} - synchronized.csv not found")

    # Convert to DataFrame with all samples (including missing ones)
    all_metrics = [metrics_dict[sample] for sample in all_expected_samples]
    results_df = pd.DataFrame(all_metrics)

    # Save to CSV
    results_df.to_csv(output_csv, index=False, float_format='%.4f')

    print(f"\n{'=' * 70}")
    print(f"Summary:")
    print(f"  Total samples expected: {len(all_expected_samples)}")
    print(f"  Samples processed: {processed_count}")
    print(f"  Missing samples: {len(all_expected_samples) - processed_count}")
    print(f"  Samples reaching 39mm: {results_df['Time_to_39mm_Gape_s'].notna().sum()}")
    print(f"\nResults saved to: {output_csv}")
    print(f"{'=' * 70}")

    # Display first few rows
    print("\nPreview (first 10 rows):")
    print(results_df.head(10).to_string(index=False))

    # Display statistics (only for non-null values)
    print("\n\nStatistics (excluding missing samples):")
    print(results_df.describe().to_string())


def create_excel_output(csv_path):
    """
    Create formatted Excel version of the CSV file.
    
    Args:
        csv_path: Path to CSV file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Create Excel file
        excel_path = csv_path.with_suffix('.xlsx')
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Batch Metrics')
            
            # Get worksheet
            ws = writer.sheets['Batch Metrics']
            
            # Format header
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 12  # Sample
            ws.column_dimensions['B'].width = 22  # Time to 39mm Gape
            ws.column_dimensions['C'].width = 22  # Force at 39mm Gape
            ws.column_dimensions['D'].width = 20  # Time to Failure
            ws.column_dimensions['E'].width = 22  # Gape at Failure
            
            # Center align numeric columns
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column > 1:  # Skip Sample column
                        cell.alignment = Alignment(horizontal='right')
        
        print(f"\nExcel file created: {excel_path}")
        return True
        
    except ImportError:
        print("\nNote: openpyxl not installed. Excel file not created.")
        print("Install with: pip install openpyxl")
        return False


if __name__ == "__main__":
    # Run main collection
    main()
    
    # Try to create Excel version
    csv_path = Path("results/batch/batch_metrics_summary.csv")
    if csv_path.exists():
        create_excel_output(csv_path)
